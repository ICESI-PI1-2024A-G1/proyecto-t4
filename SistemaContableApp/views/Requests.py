from django.conf import settings 
from SistemaContableApp.models import  *
from SistemaContableApp.forms import * 
from django.shortcuts import render, redirect
from django.core.mail import  EmailMessage 
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill



email = "pinedapablo6718@gmail.com"
email2 = "daniela32156@hotmail.com"

    
def createChargeAccountForm(request):

    if request.method == 'POST':
        form = ChargeAccountForm(request.POST, request.FILES)
        if form.is_valid():
            chargeAccount = form.save()
            sendChargeAccountFormAsExcel(request, chargeAccount)
            return redirect('viewChargeAccountForm')
    else:
        form = ChargeAccountForm()

    return render(request, 'chargeAccountForm.html', {'form': form})



def createRequisitionForm(request):

    if request.method == 'POST':
        form = RequisitionForm(request.POST)
        if form.is_valid():
            requisition = form.save()
            sendRequisitionFormAsExcel(request, requisition)
            return redirect('viewRequisitionForm')
    else:
        form = RequisitionForm()

    return render(request, 'requisitionForm.html', {'form': form})


def createExteriorPaymentForm(request):

    if request.method == 'POST':
        form = ExteriorPaymentForm(request.POST)
        if form.is_valid():
            exteriorPayment = form.save()
            sendExteriorPaymentFormAsExcel(request, exteriorPayment)
            return redirect('viewExteriorPaymentForm')
    else:
        form = ExteriorPaymentForm()

    return render(request, 'exteriorPaymentForm.html', {'form': form})



def createLegalizationForm(request):
    TravelExpenseFormSet = inlineformset_factory(
        Legalization, LegalizationExpense,
        form=TravelExpenseForm, extra=1, can_delete=True
    )

    if request.method == 'POST':
        solicitation_form = TravelExpensesSolicitationForm(request.POST, request.FILES)
        expense_formset = TravelExpenseFormSet(request.POST, request.FILES, instance=None)

        if solicitation_form.is_valid() and expense_formset.is_valid():
            solicitation = solicitation_form.save()

            expenses = expense_formset.save(commit=False)
            for expense in expenses:
                expense.solicitation = solicitation
                expense.save()

            # Enviar el archivo Excel al correo
            sendLegalizationFormAsExcel(request, solicitation)

            return redirect('viewLegalizationForm')
    else:
        solicitation_form = TravelExpensesSolicitationForm()
        expense_formset = TravelExpenseFormSet()

    return render(request, 'legalizationForm.html', {
        'solicitation_form': solicitation_form,
        'expense_formset': expense_formset
    })
    

def createAdvancePaymentForm(request):
    
    TravelAdvanceExpenseFormSet = inlineformset_factory(
    AdvancePayment, AdvanceExpense,
    form=TravelAdvanceExpenseForm, extra=1, can_delete=True,
)

    if request.method == 'POST':
        solicitation_form = TravelAdvanceSolicitationForm(request.POST, request.FILES)
        expense_formset = TravelAdvanceExpenseFormSet(request.POST, request.FILES, instance=None)

        if solicitation_form.is_valid() and expense_formset.is_valid():
            solicitation = solicitation_form.save()

            expenses = expense_formset.save(commit=False)
            for expense in expenses:
                expense.solicitation = solicitation
                expense.save()

            # Enviar el archivo Excel al correo
            sendAdvancePaymentFormAsExcel(request, solicitation)

            return redirect('viewAdvancePaymentForm')
    else:
        solicitation_form =TravelAdvanceSolicitationForm()
        expense_formset = TravelAdvanceExpenseFormSet()

    return render(request, 'advancePaymentForm.html', {
        'solicitation_form': solicitation_form,
        'expense_formset': expense_formset
    })
    



def sendChargeAccountFormAsExcel(request, chargeAccount):
    excel_filename = generateExcelChargeAccount(chargeAccount)

    # Enviar correo electrónico con el archivo Excel adjunto
    email = EmailMessage(
        'Solicitud de cuenta de cobro',
        'Adjunto se encuentra la solicitud de cuenta de cobro en formato Excel.\n Universidad Icesi Nit. 890.316.745-5.',
        settings.DEFAULT_FROM_EMAIL,
        to=["pinedapablo6718@gmail.com"]
    )
    email.attach_file(excel_filename)
    

    try:
        email.send()
        messages.success(request, 'La solicitud de cuenta de cobro e envió correctamente.')
    except Exception as e:
        messages.error(request, 'Error al enviar la solicitud de cuenta de cobro.')


def sendExteriorPaymentFormAsExcel(request, exteriorPayment):
    excel_filename = generateExcelExteriorPayment(exteriorPayment)

    # Enviar correo electrónico con el archivo Excel adjunto
    email = EmailMessage(
        'Solicitud de pago al exterior',
        'Adjunto se encuentra la solicitud de pago al exterior en formato Excel.\n Universidad Icesi Nit. 890.316.745-5.',
        settings.DEFAULT_FROM_EMAIL,
        to=["pinedapablo6718@gmail.com"]
    )
    email.attach_file(excel_filename)
    

    try:
        email.send()
        messages.success(request, 'La solicitud de pago al exterior se envió correctamente.')
    except Exception as e:
        messages.error(request, 'Error al enviar la solicitud de pago al exterior.')



def sendRequisitionFormAsExcel(request, requisition):
    excel_filename = generateExcelRequisition(requisition)

    # Enviar correo electrónico con el archivo Excel adjunto
    email = EmailMessage(
        'Solicitud de requisición',
        'Adjunto se encuentra la solicitud de requisición en formato Excel.\n Universidad Icesi Nit. 890.316.745-5.',
        settings.DEFAULT_FROM_EMAIL,
        to=["pinedapablo6718@gmail.com"]
    )
    email.attach_file(excel_filename)
    

    try:
        email.send()
        messages.success(request, 'La solicitud de requisición se envió correctamente.')
    except Exception as e:
        messages.error(request, 'Error al enviar la solicitud de requisición.')
    

def sendLegalizationFormAsExcel(request, solicitation):
    """
    Function to send the travel expenses solicitation as an Excel file.
    
    Args:
        request (HttpRequest): HTTP request object.
        solicitation (TravelExpensesSolicitation): Travel expenses solicitation instance.
    
    Returns:
        None
    """
    # Generar archivo Excel
    excel_filename = generateExcelLegalization(solicitation)

    # Enviar correo electrónico con el archivo Excel adjunto
    email = EmailMessage(
        'Solicitud de gastos de viaje',
        'Adjunto se encuentra la solicitud de gastos de viaje en formato Excel.\n Universidad Icesi Nit. 890.316.745-5.',
        settings.DEFAULT_FROM_EMAIL,
        to=["pinedapablo6718@gmail.com"]
    )
    email.attach_file(excel_filename)
    
    # Adjuntar archivos de cada gasto
    for expense in solicitation.expenses.all():
        if expense.support:
            email.attach_file(expense.support.path)

    try:
        email.send()
        messages.success(request, 'La solicitud de gastos de viaje se envió correctamente.')
    except Exception as e:
        messages.error(request, 'Error al enviar la solicitud de gastos de viaje.')
        
        
        
def sendAdvancePaymentFormAsExcel(request, solicitation) :
    """
    Function to send the travel expenses solicitation as an Excel file.
    
    Args:
        request (HttpRequest): HTTP request object.
        solicitation (TravelExpensesSolicitation): Travel expenses solicitation instance.
    
    Returns:
        None
    """
    # Generar archivo Excel
    excel_filename = generateExcelAdvancePayment(solicitation)

    # Enviar correo electrónico con el archivo Excel adjunto
    email = EmailMessage(
        'Solicitud de gastos de viaje',
        'Adjunto se encuentra la solicitud de anticipo para gastos de viaje en formato Excel.\n Universidad Icesi Nit. 890.316.745-5.',
        settings.DEFAULT_FROM_EMAIL,
        to=["pinedapablo6718@gmail.com"]
    )
    email.attach_file(excel_filename)
    

    try:
        email.send()
        messages.success(request, 'La solicitud de anticipo para gastos de viaje se envió correctamente.')
    except Exception as e:
        messages.error(request, 'Error al enviar la solicitud de anticipo para gastos de viaje.')





def generateExcelChargeAccount(chargeAccount):
    workbook = Workbook()
    worksheet = workbook.active
    
    #estilo de los bordes de la casilla
    bold_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    #borde derecho de la solicitud       
    right_border = Border(
        right=Side(style='thin')   
    )
    
    bottom_border = Border(
        bottom=Side(style='thin')
    )
    
    top_border = Border(
        top=Side(style='thin')
    )
    
    
    
    # Ancho de la columna
    column_letter = 'A'
    column_width = 33
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'B'
    column_width = 22
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'C'
    column_width = 3
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'D'
    column_width = 18
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'E'
    column_width = 31
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'F'
    column_width = 3
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'G'
    column_width = 33
    worksheet.column_dimensions[column_letter].width = column_width
    

    start_row = 1
    end_row = 64
    
    for row in range(start_row, end_row + 1):
        for col in range(7, 7 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border


    # Agregar una imagen al archivo Excel
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    #img_width, img_height = img.width, img.height


    # Escribir encabezados
    
    # Agregar Logo de icesi en la primera casilla
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    worksheet.merge_cells('A1:A4')
    merged_cell = worksheet['A1']
    #merged_cell.add_image(img,'A1' )
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('B1:E2')
    merged_cell = worksheet['B1']
    merged_cell.value = 'CONTABILIDAD'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('B3:E4')
    merged_cell = worksheet['B3']
    merged_cell.value = 'FORMATO DE GESTION DE COBRO PARA CONTRATOS DE\nGESTION HUMANA'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('F1:G2')
    merged_cell = worksheet['F1']
    merged_cell.value = 'Código:\nCTA-FR-015'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('F3:G4')
    merged_cell = worksheet['F3']
    merged_cell.value = 'Versión:\n1.0'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    

# -------------------------------------------parte de arriba del formato---------------------------------
    
    worksheet.merge_cells('B6:E6')
    merged_cell = worksheet['B6']
    merged_cell.value = 'UNIVERSIDAD ICESI'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    
    worksheet.merge_cells('B7:E7')
    merged_cell = worksheet['B7']
    merged_cell.value = 'NIT. 890.316.745-5'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    
    worksheet.merge_cells('B9:E9')
    merged_cell = worksheet['B9']
    merged_cell.value = 'DEBE A:'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    

    worksheet.merge_cells('B12:E12')
    merged_cell = worksheet['B12']
    merged_cell.value = 'Nombre:'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    merged_cell.border = top_border
    
    worksheet.merge_cells('B11:E11')
    merged_cell = worksheet['B11']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.value = chargeAccount.name
    
    worksheet.merge_cells('B15:E15')
    merged_cell = worksheet['B15']
    merged_cell.value = 'identificación:'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    merged_cell.border = top_border
    worksheet.merge_cells('B14:E14')
    merged_cell = worksheet['B14']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.value = chargeAccount.identification
    
    worksheet['B17'] = 'La suma de:'
    worksheet['B17'].font = bold_font
    
    worksheet.merge_cells('B20:C20')
    merged_cell = worksheet['B20']
    merged_cell.value = '(valor en letras)'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    merged_cell.border = top_border
    
    worksheet.merge_cells('B19:C19')
    merged_cell = worksheet['B19']
    merged_cell.value = chargeAccount.value_letters
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    worksheet['E20'] = '(valor en números)'
    worksheet['E20'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E20'].font = bold_font
    worksheet['E20'].border = top_border
    worksheet['E19'] = chargeAccount.value_numbers
    worksheet['E19'].alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B22'] = 'Concepto'
    worksheet['B22'].font = bold_font
    worksheet.merge_cells('B23:F24')
    merged_cell = worksheet['B23']
    merged_cell.value = chargeAccount.concept
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    
#------------------------ parte del formato media--------------------------------------------------

    
    worksheet.merge_cells('B28:E31')
    merged_cell = worksheet['B28']
    merged_cell.value = 'Con el fin de atender lo establecido en la Ley 2277 del año 2022, reglamentada con el Decreto 2231 de\ndiciembre de 2023 declaro, bajo la gravedad de juramento, que:\n\nLos ingresos recibidos como persona natural por concepto de renta de trabajo no proveniente de una\nrelación laboral, legal o reglamentaria correspondiente a los servicios prestados a LA UNIVERSIDAD\nICESI durante el año 2024, se les dará el siguiente tratamiento:'

    worksheet.merge_cells('B34:E35')
    merged_cell = worksheet['B34']
    merged_cell.value="a. Se tomarán costos y deducciones asociados a las rentas de trabajo por los servicios prestados:\nRetención en la fuente artículos 392 y 401 del Estatuto Tributario\nRetención en la fuente artículos 392 y 401 del Estatuto Tributario"
    
    worksheet['B37'] = 'SI'
    worksheet['B37'].font = bold_font
    
    worksheet['B38'] = 'NO'
    worksheet['B38'].font = bold_font
    
    if(chargeAccount.retentions==1):
        worksheet['C37'] = 'x'
    elif(chargeAccount.retentions==0):
        worksheet['C38'] = 'x'
        
    worksheet.merge_cells('B41:D41')
    merged_cell = worksheet['B41']
    merged_cell.value = 'b. Soy declarante del Impuesto de Renta'
    
    worksheet['E41'] = 'SI'
    worksheet['E41'].font = bold_font
    
    worksheet['E42'] = 'NO'
    worksheet['E42'].font = bold_font
    
    if(chargeAccount.declarant==1):
        worksheet['F41'] = 'x'
    elif(chargeAccount.declarant==0):
        worksheet['F42'] = 'x'
    
    worksheet.merge_cells('B44:D44')
    merged_cell = worksheet['B44']
    merged_cell.value = 'c. Soy residente fiscal en Colombia'
    
    worksheet['E44'] = 'SI'
    worksheet['E44'].font = bold_font
    
    worksheet['E45'] = 'NO'
    worksheet['E45'].font = bold_font
    
    if(chargeAccount.colombian_resident==1):
        worksheet['F44'] = 'x'
    elif(chargeAccount.colombian_resident==0):
        worksheet['F45'] = 'x'
        
        
    worksheet['B48'] = 'Ciudad y Fecha'
    worksheet['B48'].font = bold_font
    worksheet.merge_cells('C48:D48')
    merged_cell = worksheet['C48']
    merged_cell.value = chargeAccount.city 
    merged_cell.border = bottom_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['E48'].value = chargeAccount.date
    worksheet['E48'].border = bottom_border
    worksheet['E48'].alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B51'] = 'Firma'
    worksheet['B51'].font = bold_font
    worksheet.merge_cells('C51:E51')
    merged_cell = worksheet['C51']
    merged_cell.value = chargeAccount.name
    merged_cell.border = bottom_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B52'] = 'C.C'
    worksheet['B52'].font = bold_font
    worksheet.merge_cells('C52:E52')
    merged_cell = worksheet['C52']
    merged_cell.value = chargeAccount.identification
    merged_cell.border = bottom_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B53'] = 'Dirección'
    worksheet['B53'].font = bold_font
    worksheet.merge_cells('C53:E53')
    merged_cell = worksheet['C53']
    merged_cell.value = chargeAccount.addres
    merged_cell.border = bottom_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B54'] = 'Telefono'
    worksheet['B54'].font = bold_font
    worksheet.merge_cells('C54:E54')
    merged_cell = worksheet['C54']
    merged_cell.value = chargeAccount.phone
    merged_cell.border = bottom_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    

#------------------------ parte del formato baja--------------------------------------------------


    worksheet.merge_cells('B57:C57')
    merged_cell = worksheet['B57']
    merged_cell.value = "Nombre del banco"
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet.merge_cells('B58:C58')
    merged_cell = worksheet['B58']
    merged_cell.value = chargeAccount.bank
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    worksheet['D57'].value = "Tipo de cuenta"
    worksheet['D57'].border = thin_border
    worksheet['D57'].font = bold_font
    worksheet['D57'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['D58'] = chargeAccount.type
    worksheet['D58'].border = thin_border
    worksheet['D58'].alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['E57'].value = "No."
    worksheet['E57'].border = thin_border
    worksheet['E57'].font = bold_font
    worksheet['E57'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E58'] = chargeAccount.account_number
    worksheet['E58'].border = thin_border
    worksheet['E58'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('B60:E60')
    merged_cell = worksheet['B60']
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.value = "Cex No."
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    worksheet.merge_cells('B61:E61')
    merged_cell = worksheet['B61']
    merged_cell.value = chargeAccount.cex
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    filename = f'media/chargeAccount_{chargeAccount.id}.xlsx'
    workbook.save(filename)
    return filename


def generateExcelExteriorPayment(exteriorPayment):
    workbook = Workbook()
    worksheet = workbook.active
    
    #borde derecho de la solicitud       
    right_border = Border(
        right=Side(style='thin')   
    )
    start_row = 1
    end_row = 36
    
    for row in range(start_row, end_row + 1):
        for col in range(1, 1 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border
    
    
    
    top_border = Border(
        top=Side(style='thin')
    )

    
    
    column_letter = 'A'
    column_width = 80
    worksheet.column_dimensions[column_letter].width = column_width
    
    bold_font = Font(bold=True)
    
    worksheet['A2'] = 'SOLICITUD DE PAGO AL EXTERIOR'
    worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A2'].font = bold_font
    worksheet['A2'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    worksheet['A4'] = 'DATOS DEL BENEFICIARIO'
    worksheet['A4'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A4'].font = bold_font
    worksheet['A4'].fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    
    worksheet['A5'] = 'Nombre Completo'
    worksheet['A5'].font = bold_font
    worksheet['A6'] = exteriorPayment.beneficiary_name + " " + exteriorPayment.beneficiary_last_name
    
    worksheet['A7'] = 'Tipo de identificación'
    worksheet['A7'].font = bold_font
    worksheet['A8'] = exteriorPayment.beneficiary_document_type
    
    worksheet['A9'] = 'Numero de identificación'
    worksheet['A9'].font = bold_font
    worksheet['A10'] = exteriorPayment.beneficiary_document_no
    
    worksheet['A11'] = 'Numero de Pasaporte:'
    worksheet['A11'].font = bold_font
    worksheet['A12'] = exteriorPayment.passport_number
    
    worksheet['A13'] = 'ciudad de expedición:'
    worksheet['A13'].font = bold_font
    worksheet['A14'] = exteriorPayment.passport_expedition_city
    
    worksheet['A15'] = 'Datos de domicilio'
    worksheet['A15'].font = bold_font
    worksheet['A16'] = exteriorPayment.address
    
    
    worksheet['A19'] = 'DATOS DE LA ENTIDAD BANCARIA DEL BENEFICIARIO'
    worksheet['A19'].font = bold_font
    worksheet['A19'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A19'].fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    
    worksheet['A21'] = 'Nombre de la entidad bancaria:'
    worksheet['A21'].font = bold_font
    worksheet['A22'] = exteriorPayment.bank_name
    
    worksheet['A23'] = 'Tipo de cuenta(Ahorros o Corriente):'
    worksheet['A23'].font = bold_font
    worksheet['A24'] = exteriorPayment.account_type
    
    worksheet['A25'] = 'Código Swift:'
    worksheet['A25'].font = bold_font
    worksheet['A26'] = exteriorPayment.swift_code
    
    worksheet['A27'] = 'Tipo de código IBAN/ABA'
    worksheet['A27'].font = bold_font
    worksheet['A28'] = exteriorPayment.iban_aba_code_type
    
    worksheet['A29'] = 'Código IBAN/ABA'
    worksheet['A29'].font = bold_font
    worksheet['A30'] = exteriorPayment.iban_aba_code
    
    worksheet['A31'] = 'Nombre de la cuenta:'
    worksheet['A31'].font = bold_font
    worksheet['A32'] = exteriorPayment.account_name
    
    worksheet['A33'] = 'Numero de la cuenta:'
    worksheet['A33'].font = bold_font
    worksheet['A34'] = exteriorPayment.account_number
    
    worksheet['A35'] = 'Dirección de la entidad bancaria'
    worksheet['A35'].font = bold_font
    worksheet['A36'] = exteriorPayment.bank_address
    
    
    filename = f'media/exteriorPayment_{exteriorPayment.id}.xlsx'
    workbook.save(filename)
    return filename




def generateExcelRequisition(requisition):
    workbook = Workbook()
    worksheet = workbook.active
    
    #estilo de los bordes de la casilla
    bold_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    #borde derecho de la solicitud       
    right_border = Border(
        right=Side(style='thin')   
    )
    
    bottom_border = Border(
        bottom=Side(style='thin')
    )
    
    top_border = Border(
        top=Side(style='thin')
    )
    
    
    
    # Ancho de la columna
    column_letter = 'A'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'B'
    column_width = 20
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'C'
    column_width = 20
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'D'
    column_width = 2
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'E'
    column_width = 23
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'F'
    column_width = 2
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'G'
    column_width = 33
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'H'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    

    start_row = 1
    end_row = 42
    
    for row in range(start_row, end_row + 1):
        for col in range(8, 8 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border


    # Agregar una imagen al archivo Excel
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    #img_width, img_height = img.width, img.height


    # Escribir encabezados
    
    # Agregar Logo de icesi en la primera casilla
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    worksheet.merge_cells('A1:B4')
    merged_cell = worksheet['A1']
    #merged_cell.add_image(img,'A1' )
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('C1:F2')
    merged_cell = worksheet['C1']
    merged_cell.value = 'CONTABILIDAD'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('C3:F4')
    merged_cell = worksheet['C3']
    merged_cell.value = 'FORMATO DE REQUISICIÓN'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('G1:H2')
    merged_cell = worksheet['G1']
    merged_cell.value = 'Código:\nCTA-FR-005'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('G3:H4')
    merged_cell = worksheet['G3']
    merged_cell.value = 'Versión:\n1.0'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    

# -------------------------------------------parte de arriba del formato---------------------------------
    worksheet['B6'] = 'No. de Radicado'
    worksheet['B6'].font = bold_font
    worksheet['C6'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    worksheet['C6'] = requisition.radicate
    worksheet['C6'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet['E6'] = 'Código de Orden de pago'
    worksheet['E6'].font = bold_font
    worksheet.merge_cells('F6:G6')
    merged_cell = worksheet['F6']
    merged_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    worksheet['F6'] = requisition.payment_order_code
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B8'] = 'Fecha'
    worksheet['B8'].font = bold_font
    worksheet.merge_cells('C8:F8')
    merged_cell = worksheet['C8']
    worksheet['C8'] = requisition.date
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border
    
       
    worksheet['B9'] = 'A favor de'
    worksheet['B9'].font = bold_font
    worksheet.merge_cells('C9:F9')
    merged_cell = worksheet['C9']
    worksheet['C9'] = requisition.beneficiaryName
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border
    
    worksheet['B10'] = 'Numero de cedula'
    worksheet['B10'].font = bold_font
    worksheet.merge_cells('C10:F10')
    merged_cell = worksheet['C10']
    worksheet['C10'] = requisition.idNumber
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border
    
    worksheet['B11'] = 'Cargo'
    worksheet['B11'].font = bold_font
    worksheet.merge_cells('C11:F11')
    merged_cell = worksheet['C11']
    worksheet['C11'] = requisition.charge
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border

    worksheet['B12'] = 'Dependencia'
    worksheet['B12'].font = bold_font
    worksheet.merge_cells('C12:F12')
    merged_cell = worksheet['C12']
    worksheet['C12'] = requisition.dependency
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border

    worksheet['B13'] = 'Cenco'
    worksheet['B13'].font = bold_font
    worksheet.merge_cells('C13:F13')
    merged_cell = worksheet['C13']
    worksheet['C13'] = requisition.cenco
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border

    worksheet['B14'] = 'Valor'
    worksheet['B14'].font = bold_font
    worksheet.merge_cells('C14:F14')
    merged_cell = worksheet['C14']
    worksheet['C14'] = requisition.value
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border

    worksheet['B15'] = 'Concepto'
    worksheet['B15'].font = bold_font
    worksheet.merge_cells('C15:F15')
    merged_cell = worksheet['C15']
    worksheet['C15'] = requisition.concept
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = top_border
    
    worksheet.merge_cells('C16:F16')
    merged_cell = worksheet['C16']
    merged_cell.border = top_border
    
    
    #Gris el valor de los datos
    start_row = 8
    end_row = 15
    start_column = 3  # Columna A
    end_column = 6  # Columna H
    
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
            cell.border = thin_border
    row = 21
    
#------------------------ parte del formato media (Descripcion)--------------------------------------------------

    
    worksheet.merge_cells('B17:G17')
    merged_cell = worksheet['B17']
    merged_cell.border = bottom_border
    worksheet['B17'] = 'DESCRIPCIÓN'
    worksheet['B17'].font = bold_font
    worksheet['B17'].alignment = Alignment(horizontal='center', vertical='center')
    
    
    worksheet.merge_cells('B18:G26')
    merged_cell = worksheet['B18']
    worksheet['B18'] = requisition.description
    
    
    
    
    #borde derecho de la Descripción    
    start_row = 18
    end_row = 26
    
    for row in range(start_row, end_row + 1):
        for col in range(7, 7 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border
            
      
    #borde derecho de la Descripción    
    start_row = 18
    end_row = 26
    
    for row in range(start_row, end_row + 1):
        for col in range(1, 1 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border      
            
    #Gris el valor de la Descripcion
    start_row = 18
    end_row = 26
    start_column = 2  # Columna A
    end_column = 7  # Columna H
    
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
            cell.border = thin_border
    row = 21

#----------------------------------Parte de abajo del formato-------------------------------------------------------------------------

    worksheet['B28'] = 'Datos para pago: '
    worksheet['B28'].font = bold_font
    
    worksheet.merge_cells('C28:E28')
    merged_cell = worksheet['C28']
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['C28'] = requisition.paymentMethod
    
    worksheet['B30'] = 'Ahorros'
    worksheet['B30'].font = bold_font
    
    worksheet['B31'] = 'Corriente'
    worksheet['B31'].font = bold_font
    
    if(requisition.typeAccount=="De ahorros"):
        worksheet['C30'] = 'x'
    elif(requisition.typeAccount=="Corriente"):
        worksheet['C31'] = 'x'
        
        
    worksheet['E30'] = '# Cuenta Bancaria'
    worksheet['E30'].font = bold_font
    
    worksheet['G30'] = requisition.account_number
    worksheet['G30'].font = bold_font
    worksheet['G30'].fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    worksheet['G30'].alignment = Alignment(horizontal='center', vertical='center')
        
    worksheet['B35'] = 'Firma Ordenador de Gasto'
    worksheet['B35'].border = top_border
    worksheet['B35'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B34'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B34'] = requisition.beneficiaryName
    
    worksheet['E35'] = 'Firma de quien elabora'
    worksheet['E35'].border = top_border
    worksheet['E35'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E34'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E34'] = requisition.authorName

    worksheet['B38'] = 'Nombre Ordenador de gasto'
    worksheet['B38'].border = top_border
    worksheet['B38'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B37'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B37'] = requisition.beneficiaryName
    
    worksheet['E38'] = 'Nombre de quien elabora'
    worksheet['E38'].border = top_border
    worksheet['E38'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E37'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E37'] = requisition.authorName
    
    
    worksheet.merge_cells('B43:G43')
    merged_cell = worksheet['B43']
    merged_cell.value = 'Campo de uso exclusivo de la Oficina de Contabilidad'
    merged_cell.font = bold_font  
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet.merge_cells('B44:G44')
    merged_cell = worksheet['B44']
    merged_cell.value = 'Motivo de devolución:'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    
    worksheet.merge_cells('A46:H46')
    merged_cell = worksheet['A46']
    merged_cell.value = 'Espacio para ser diligenciado por la oficina de contabilidad:'
    merged_cell.font = bold_font  
    merged_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')


    
    
    filename = f'media/requisition_{requisition.id}.xlsx'
    workbook.save(filename)
    return filename
    
    
def generateExcelAdvancePayment(solicitation):
    workbook = Workbook()
    worksheet = workbook.active
    
    #estilo de los bordes de la casilla
    bold_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    
    
    # Ancho de la columna
    column_letter = 'A'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'B'
    column_width = 20
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'C'
    column_width = 20
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'D'
    column_width = 2
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'E'
    column_width = 23
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'F'
    column_width = 2
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'G'
    column_width = 33
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'H'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    
    #borde derecho de la solicitud       
    right_border = Border(
        right=Side(style='thin')   
    )
    start_row = 1
    end_row = 42
    
    for row in range(start_row, end_row + 1):
        for col in range(8, 8 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border


    # Agregar una imagen al archivo Excel
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    #img_width, img_height = img.width, img.height


    # Escribir encabezados
    
    # Agregar Logo de icesi en la primera casilla
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    worksheet.merge_cells('A1:B4')
    merged_cell = worksheet['A1']
    #merged_cell.add_image(img,'A1' )
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('C1:F2')
    merged_cell = worksheet['C1']
    merged_cell.value = 'CONTABILIDAD'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('C3:F4')
    merged_cell = worksheet['C3']
    merged_cell.value = 'FORMATO SOLICITUD DE ANTICIPO PARA GASTOS DE VIAJE'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('G1:H2')
    merged_cell = worksheet['G1']
    merged_cell.value = 'Código:\nCTA-FR-008'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('G3:H4')
    merged_cell = worksheet['G3']
    merged_cell.value = 'Versión:\n1.0'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    

# -------------------------------------------parte de arriba del formato---------------------------------
    
    worksheet['B6'] = 'No. de Radicado'
    worksheet['B6'].font = bold_font
    worksheet['C6'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    worksheet['C6'] = solicitation.radicate
    worksheet['C6'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet['E6'] = 'Código de Orden de pago'
    worksheet['E6'].font = bold_font
    worksheet.merge_cells('F6:G6')
    merged_cell = worksheet['F6']
    merged_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    worksheet['F6'] = solicitation.payment_order_code
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B8'] = 'Fecha de solicitud'
    worksheet['B8'].font = bold_font
    worksheet['C8'] = solicitation.request_date
    worksheet['C8'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet['B9'] = 'Nombre del viajero'
    worksheet['B9'].font = bold_font
    worksheet.merge_cells('C9:G9')
    merged_cell = worksheet['C9']
    worksheet['C9'] = solicitation.traveler_name
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B10'] = 'No. de identificación'
    worksheet['B10'].font = bold_font
    worksheet.merge_cells('C10:D10')
    merged_cell = worksheet['C14']
    worksheet['C10'] = solicitation.traveler_id
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['E10'] = 'Centro de Costos'
    worksheet['E10'].font = bold_font
    worksheet.merge_cells('F10:G10')
    merged_cell = worksheet['F10']
    worksheet['F10'] = solicitation.cost_center
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet['B11'] = 'Dependencia'
    worksheet['B11'].font = bold_font
    worksheet.merge_cells('C11:G11')
    merged_cell = worksheet['C11']
    worksheet['C11'] = solicitation.dependency
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet['B13'] = 'Ciudad de destino'
    worksheet['B13'].font = bold_font
    worksheet.merge_cells('C13:D13')
    merged_cell = worksheet['C13']
    worksheet['C13'] = solicitation.destiny_city
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet['B14'] = 'Fecha de Salida'
    worksheet['B14'].font = bold_font
    worksheet.merge_cells('C14:D14')
    merged_cell = worksheet['C14']
    worksheet['C14'] = solicitation.travel_date
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet['E14'] = 'Fecha de Regreso'
    worksheet['E14'].font = bold_font
    worksheet.merge_cells('F14:G14')
    merged_cell = worksheet['F14']
    worksheet['F14'] = solicitation.return_date
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet['B15'] = 'Motivo del Viaje'
    worksheet['B15'].font = bold_font
    worksheet.merge_cells('C15:G15')
    merged_cell = worksheet['C15']
    worksheet['C15'] = solicitation.motive
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet['B17'] = 'Tipo de moneda:'
    worksheet['B17'].font = bold_font
    worksheet.merge_cells('C17:G17')
    merged_cell = worksheet['C17']
    worksheet['C17'] = solicitation.currency_type_of_advance_value
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet.merge_cells('B18:D18')
    merged_cell = worksheet['B18']
    worksheet['B18'].value = 'Indique el último día que estará en Icesi antes de su viaje'
    worksheet.merge_cells('E18:G18')
    merged_cell = worksheet['E18']
    worksheet['E18'] = solicitation.last_day_in_icesi
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
#------------------------ parte del formato media (tabla de gastos)--------------------------------------------------
    worksheet.merge_cells('C20:E20')
    merged_cell = worksheet['C20']
    worksheet['C20'] = 'PRESUPUESTO DE GASTOS'
    worksheet['C20'].font = bold_font
    worksheet['C20'].alignment = Alignment(horizontal='center', vertical='center')
    
    
    #Gris el valor de los gastos
    start_row = 21
    end_row = 29
    start_column = 5  # Columna A
    end_column = 5  # Columna H
    
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
            cell.border = thin_border
    row = 21
    for expense in solicitation.expenses.all():
        worksheet.cell(row=row, column=3, value=expense.category)
        worksheet.cell(row=row, column=5, value=expense.money_value)

        row += 1
        
    worksheet['C29'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['C29'] = 'TOTAL DEL ANTICIPO'
    worksheet['C29'].font = bold_font
    worksheet['E29'].value ='=SUM(E21:E28)'
    worksheet['E29'].border = thin_border


    
    
#----------------------------------Parte de abajo del formato-------------------------------------------------------------------------
    worksheet.merge_cells('B30:G31')
    merged_cell = worksheet['B30']
    merged_cell.value = 'Autorización de descuento : Si pasados 15 días después de finalizar el viaje no he legalizado este anticipo;\n autorizo que su valor me sea descontado por nómina en lel mes más próximo.'
    
    
    top_border = Border(
        top=Side(style='thin')
    )
    
    worksheet['B34'] = 'Firma del viajero'
    worksheet['B34'].border = top_border
    worksheet['B34'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B33'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B33'] = solicitation.traveler_name
    
    worksheet['E34'] = 'Firma Ordenador de Gasto'
    worksheet['E34'].border = top_border
    worksheet['E34'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E33'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E33'] = solicitation.orderer_name

    worksheet['B37'] = 'Nombre de quien elabora'
    worksheet['B37'].border = top_border
    worksheet['B37'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B36'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['B36'] = solicitation.orderer_name
    
    worksheet['E37'] = 'Nombre Ordenador de Gasto'
    worksheet['E37'].border = top_border
    worksheet['E37'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E36'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E36'] = solicitation.orderer_name
    
    
    worksheet.merge_cells('B39:G39')
    merged_cell = worksheet['B39']
    merged_cell.value = 'Campo de uso exclusivo de la Oficina de Contabilidad'
    merged_cell.font = bold_font  
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    
    worksheet.merge_cells('B40:G40')
    merged_cell = worksheet['B40']
    merged_cell.value = 'Motivo de devolución:'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    
    
    
    worksheet.merge_cells('A42:H42')
    merged_cell = worksheet['A42']
    merged_cell.value = 'Espacio para ser diligenciado por la oficina de contabilidad:'
    merged_cell.font = bold_font  
    merged_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    
    # Guardar el archivo Excel
    filename = f'media/advancePayment_{solicitation.id}.xlsx'
    workbook.save(filename)
    return filename


 



def generateExcelLegalization(solicitation):
    workbook = Workbook()
    worksheet = workbook.active
    
    
    
    #estilo de los bordes de la casilla
    bold_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    #Aplicarle bordes de las casillas de la tabla de gastos
    start_row = 20
    end_row = 34
    start_column = 1  # Columna A
    end_column = 9  # Columna H
    
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            
     
    #borde derecho de la solicitud       
    right_border = Border(
        right=Side(style='thin')   
    )
    start_row = 1
    end_row = 48
    
    for row in range(start_row, end_row + 1):
        for col in range(9, 9 + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = right_border
            
            
    
    # Ancho de la columna
    column_letter = 'A'
    column_width = 32
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'B'
    column_width = 9
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'C'
    column_width = 18
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'D'
    column_width = 14
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'E'
    column_width = 38
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'F'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'G'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'H'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width
    
    column_letter = 'I'
    column_width = 11
    worksheet.column_dimensions[column_letter].width = column_width



    # Agregar una imagen al archivo Excel
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    #img_width, img_height = img.width, img.height


    # Escribir encabezados
    
    # Agregar Logo de icesi en la primera casilla
    #image_path = 'SistemaContableApp\static\images\LogoIcesi.jpg'  # Reemplaza con la ruta de tu imagen
    #img = Image(image_path)
    worksheet.merge_cells('A1:A4')
    merged_cell = worksheet['A1']
    #merged_cell.add_image(img,'A1' )
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('B1:G2')
    merged_cell = worksheet['B1']
    merged_cell.value = 'CONTABILIDAD'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('B3:G4')
    merged_cell = worksheet['B3']
    merged_cell.value = 'FORMATO DE LEGALIZACIÓN DE GASTOS DE VIAJE'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('H1:I2')
    merged_cell = worksheet['H1']
    merged_cell.value = 'Código:\nCTA-FR-008'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    worksheet.merge_cells('H3:I4')
    merged_cell = worksheet['H3']
    merged_cell.value = 'Versión:\n1.0'
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.border = thin_border
    merged_cell.font = bold_font
    
    

    
    
    # -------------------------------------------parte de arriba del formato---------------------------------
    worksheet['A6'] = 'Fecha de legalización'
    worksheet['A6'].font = bold_font
    worksheet['B6'] = solicitation.legalization_date

    
    worksheet['A7'] = 'Nombre del viajero'
    worksheet['A7'].font = bold_font
    worksheet['B7'] = solicitation.traveler_name

    worksheet['A8'] = 'No. de Identificación'
    worksheet['A8'].font = bold_font
    worksheet['B8'] = solicitation.identification

    worksheet['E8'] = 'Centro de Costos'
    worksheet['E8'].font = bold_font
    worksheet['F8'] = solicitation.cost_center

    worksheet['A9'] = 'Dependencia'
    worksheet['A9'].font = bold_font
    worksheet['B9'] = solicitation.dependency

    worksheet['A11'] = 'Ciudad de destino'
    worksheet['A11'].font = bold_font
    worksheet['B11'] = solicitation.destiny_city

    worksheet['A12'] = 'Fecha de Salida'
    worksheet['A12'].font = bold_font
    worksheet['B12'] = solicitation.travel_date

    worksheet['E12'] = 'Fecha de Regreso'
    worksheet['E12'].font = bold_font
    worksheet['F12'] = solicitation.return_date

    worksheet['A13'] = 'Motivo del Viaje'
    worksheet['A13'].font = bold_font
    worksheet['B13'] = solicitation.motive
    

    #------------------------ parte del formato media (tabla de gastos)--------------------------------------------------
    worksheet['A16'] = 'Relación de gastos'
    worksheet['A16'].font = bold_font
    worksheet['A16'].fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    worksheet['A16'].alignment = Alignment(horizontal='center', vertical='center')
    
    
    worksheet.merge_cells('A18:A19')
    merged_cell = worksheet['A18']
    merged_cell.value = 'Rubro'
    merged_cell.font = bold_font
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('B18:B19')
    merged_cell = worksheet['B18']
    merged_cell.value = '# de\nsoporte'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font
    
    worksheet.merge_cells('C18:C19')
    merged_cell = worksheet['C18']
    merged_cell.value = 'Nombre del tercero\nde la factura'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font 
    
    worksheet.merge_cells('D18:D19')
    merged_cell = worksheet['D18']
    merged_cell.value = 'NIT del\ntercero\nde la factura'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font    
    
    worksheet.merge_cells('E18:E19')
    merged_cell = worksheet['E18']
    merged_cell.value = 'Concepto de la compra'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font 
    
    worksheet.merge_cells('F18:H18')
    merged_cell = worksheet['F18']
    merged_cell.value = 'Si su anticipo fue en:'
    merged_cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    merged_cell.border = thin_border
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = bold_font 
    

    worksheet['F19'] = 'Pesos\ncolombianos'
    worksheet['F19'].border = thin_border
    
    worksheet['G19'] = 'Dólares'
    worksheet['G19'].border = thin_border
    
    worksheet['H19'] = 'Euros'
    worksheet['H19'].border = thin_border

    
    
    row = 20
    for expense in solicitation.expenses.all():
        worksheet.cell(row=row, column=1, value=expense.category)
        worksheet.cell(row=row, column=2, value=expense.support_no)
        worksheet.cell(row=row, column=3, value=expense.third_person_name)
        worksheet.cell(row=row, column=4, value=expense.third_person_nit)
        worksheet.cell(row=row, column=5, value=expense.concept)

        # Diferencia el valor de money_value según el tipo de moneda
        if expense.money_type == 'PESOS COLOMBIANOS':
            worksheet.cell(row=row, column=6, value=expense.money_value)
        elif expense.money_type == 'DOLARES':
            worksheet.cell(row=row, column=7, value=expense.money_value)
        elif expense.money_type == 'EUROS':
            worksheet.cell(row=row, column=8, value=expense.money_value)

        row += 1
            

    # Total de gastos
    worksheet['A31'] = 'Total de gastos'
    worksheet['F31'].value = '=SUM(F20:F30)'  # Suma de pesos colombianos
    worksheet['G31'].value = '=SUM(G20:G30)'  # Suma de dólares
    worksheet['H31'].value = '=SUM(H20:H30)'  # Suma de euros
    
    
    #color gris para la parte de total de gastos
    start_row = 31
    end_row = 31
    start_column = 1  # Columna A
    end_column = 8  # Columna H
    
    for col in range(start_row, end_row + 1):
        for row in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
       
          
    # Valor del anticipo
    worksheet['A32'] = 'Valor del anticipo'
    if solicitation.currency_type_of_advance_value == 'PESOS COLOMBIANOS':

        worksheet['F32'] = solicitation.advance_payment_value  # Valor de anticipo en pesos colombianos
        worksheet['G32'] = 0  # Valor de anticipo en dólares
        worksheet['H32'] = 0  # Valor de anticipo en euros
        
    elif solicitation.currency_type_of_advance_value == 'DOLARES':
        
        worksheet['F32'] = 0  # Valor de anticipo en pesos colombianos
        worksheet['G32'] = solicitation.advance_payment_value  # Valor de anticipo en dólares
        worksheet['H32'] = 0  # Valor de anticipo en euros
        
    elif solicitation.currency_type_of_advance_value == 'EUROS':
        
        worksheet['F32'] = 0  # Valor de anticipo en pesos colombianos
        worksheet['G32'] = 0  # Valor de anticipo en dólares
        worksheet['H32'] = solicitation.advance_payment_value  # Valor de anticipo en euros
        
    
    #color azul para la parte de valor de anticipo
    
    
    start_row = 32
    end_row = 32
    start_column = 1  # Columna A
    end_column = 8  # Columna H
    
    for col in range(start_row, end_row + 1):
        for row in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
  

    # Saldo a favor del empleado
    worksheet['A33'] = 'Saldo a favor del empleado'
    worksheet['F33'].value = '=IF(F31>F32,F31-F32,0)'  # Saldo en pesos colombianos
    worksheet['G33'].value = '=IF(G31>G32,G31-G32,0)'  # Saldo en dólares
    worksheet['H33'].value = '=IF(H31>H32,H31-H32,0)'  # Saldo en euros

    # Saldo a favor de ICESI
    worksheet['A34'] = 'Saldo a favor de ICESI'
    worksheet['F34'].value = '=IF(F31<F32,F32-F31,0)'  # Saldo en pesos colombianos
    worksheet['G34'].value = '=IF(G31<G32,G32-G31,0)'  # Saldo en dólares
    worksheet['H34'].value = '=IF(H31<H32,H32-H31,0)'  # Saldo en euros

    

    #----------------------------------Parte de abajo del formato-------------------------------------------------------------------------
    worksheet.merge_cells('A36:E37')
    merged_cell = worksheet['A36']
    merged_cell.value = 'Autorizo que el saldo a favor de la Universidad Icesi, sea descontado en una sola\ncuota en el siguiente pago de nómina.'
    
    
    if solicitation.descount_in_one_quote == 1:
        worksheet['C38'].value = 'SÍ autorizo'
        worksheet['C38'].fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    elif solicitation.descount_in_one_quote == 0:
        worksheet['C38'].value = 'NO autorizo'
        worksheet['C38'].fill = PatternFill(start_color='FFABAB', end_color='FFABAB', fill_type='solid')
    
    top_border = Border(
        top=Side(style='thin')
    )
    
    worksheet['A41'] = 'Nombre de quien elabora'
    worksheet['A41'].border = top_border
    worksheet['A41'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A40'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A40'] = solicitation.elaborator_name
    
    worksheet['E41'] = 'Nombre Ordenador de gasto'
    worksheet['E41'].border = top_border
    worksheet['E41'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E40'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['E40'] = solicitation.orderer_name
    
    
    worksheet['A45'] = 'Banco'
    worksheet['A45'].font = bold_font  
    worksheet['B45'] = solicitation.bank
    
    worksheet['A46'] = '# de cuenta'
    worksheet['A46'].font = bold_font  
    worksheet['B46'] = solicitation.account_number
    
    worksheet['D45'] = 'Tipo de cuenta'
    worksheet['D45'].font = bold_font  
    worksheet['E45'] = solicitation.type_account
    
    
    worksheet.merge_cells('A48:I48')
    merged_cell = worksheet['A48']
    merged_cell.value = 'OBSERVACIONES:'
    merged_cell.font = bold_font  
    merged_cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    # Guardar el archivo Excel
    filename = f'media/legalizacion_{solicitation.id}.xlsx'
    workbook.save(filename)
    return filename



        
        